#!/usr/bin/env python3
"""
Transform Conway SC Leapfrog Excel data into Healthcare Dashboard format
Updated for Fall 2025 data
"""

import openpyxl
import json
import os
from datetime import datetime

# Path to Excel file
EXCEL_PATH = '/Users/heilman/Desktop/Conway SC Leapfrog Historical Measure Analysis 2023-2025 incl Fall 2025.xlsx'

def parse_excel():
    """Parse the Excel file and extract all data"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active  # Use active sheet

    measures = []

    # Parse each measure (rows 4 onwards, row 3 is headers)
    for row in range(4, 100):  # Check up to row 100
        domain = ws.cell(row, 1).value
        measure_id = ws.cell(row, 2).value
        measure_name = ws.cell(row, 3).value

        if not measure_name or measure_name == 'Measure Name':
            break

        measure = {
            'domain': domain,
            'measureId': measure_id,
            'measureName': measure_name,
            'periods': {
                '2023 Spring': {
                    'conwayResult': ws.cell(row, 4).value,
                    'conwayZScore': ws.cell(row, 5).value,
                    'nationalMean': ws.cell(row, 6).value
                },
                '2023 Fall': {
                    'conwayResult': ws.cell(row, 7).value,
                    'conwayZScore': ws.cell(row, 8).value,
                    'nationalMean': ws.cell(row, 9).value
                },
                '2024 Spring': {
                    'conwayResult': ws.cell(row, 10).value,
                    'conwayZScore': ws.cell(row, 11).value,
                    'nationalMean': ws.cell(row, 12).value
                },
                '2024 Fall': {
                    'conwayResult': ws.cell(row, 13).value,
                    'conwayZScore': ws.cell(row, 14).value,
                    'nationalMean': ws.cell(row, 15).value
                },
                '2025 Spring': {
                    'conwayResult': ws.cell(row, 16).value,
                    'conwayZScore': ws.cell(row, 17).value,
                    'nationalMean': ws.cell(row, 18).value
                },
                '2025 Fall': {
                    'conwayResult': ws.cell(row, 19).value,  # Column S
                    'conwayZScore': 0,  # No Z-score in Excel for Fall 2025
                    'nationalMean': ws.cell(row, 20).value  # Column T
                }
            },
            'summary': {
                'conwayMean': ws.cell(row, 21).value,  # Column U
                'nationalMean': ws.cell(row, 22).value,  # Column V
                'conwaySlope': ws.cell(row, 23).value,  # Column W
                'nationalSlope': ws.cell(row, 24).value,  # Column X
                'weight': ws.cell(row, 25).value  # Column Y
            },
            'forecasts': {
                '2026': {
                    'conwayResult': ws.cell(row, 26).value,  # Column Z
                    'nationalMean': ws.cell(row, 27).value  # Column AA
                },
                '2027': {
                    'conwayResult': ws.cell(row, 28).value,  # Column AB
                    'nationalMean': ws.cell(row, 29).value  # Column AC
                }
            }
        }

        measures.append(measure)

    return measures

def generate_scatter_plot_data(measures):
    """Generate scatterPlotData array for the dashboard"""
    data = []

    for measure in measures:
        data.append({
            'measureGroup': measure['domain'],
            'measureCategory': measure['domain'],
            'measureId': measure['measureId'],
            'measureName': measure['measureName'],
            'conwaySlope': measure['summary']['conwaySlope'] if measure['summary']['conwaySlope'] else 0,
            'nationalSlope': measure['summary']['nationalSlope'] if measure['summary']['nationalSlope'] else 0,
        })

    return data

def generate_performance_data(measures):
    """Generate performanceData object with time series for each measure"""
    performance_data = {}

    for measure in measures:
        measure_id = measure['measureId']
        time_series = []

        # Add historical periods
        for period, data in measure['periods'].items():
            conway = data['conwayResult']
            national = data['nationalMean']

            # Handle special values
            if conway == 'Not Included' or conway == '*':
                conway = None
            elif isinstance(conway, str) and conway.strip() == '':
                conway = None

            if national == 'Not Included' or national == '*':
                national = None
            elif isinstance(national, str) and national.strip() == '':
                national = None

            # Determine status
            if conway is None or national is None:
                status = 'no data'
            else:
                # Simple comparison - will refine based on measure type
                diff_percent = abs(conway - national) / national * 100 if national != 0 else 0

                if diff_percent <= 5:
                    status = 'similar'
                elif conway < national:
                    status = 'better'  # Generally lower is better for most outcome measures
                else:
                    status = 'worse'

            time_series.append({
                'year': period,
                'conway': conway,
                'national': national,
                'status': status
            })

        # Add forecasts
        for year, forecast_data in measure['forecasts'].items():
            conway_forecast = forecast_data['conwayResult']
            national_forecast = forecast_data['nationalMean']

            if conway_forecast is not None and national_forecast is not None:
                time_series.append({
                    'year': year,
                    'conway': conway_forecast,
                    'national': national_forecast,
                    'status': 'forecast'
                })

        # Add summary row
        time_series.append({
            'year': '2023-2025',
            'conwayMean': measure['summary']['conwayMean'],
            'nationalMean': measure['summary']['nationalMean'],
            'conwaySlope': measure['summary']['conwaySlope'],
            'nationalSlope': measure['summary']['nationalSlope'],
            'weight': measure['summary']['weight']
        })

        performance_data[measure_id] = time_series

    return performance_data

def generate_measure_data(measures):
    """Generate measureData object organized by domain"""
    measure_data = {}

    # Group measures by domain
    for measure in measures:
        domain = measure['domain']

        if domain not in measure_data:
            measure_data[domain] = []

        measure_data[domain].append({
            'id': measure['measureId'],
            'name': measure['measureName'],
            'weight': measure['summary']['weight'] if measure['summary']['weight'] else 0
        })

    return measure_data

def calculate_underperforming_measures(measures):
    """Calculate underperforming measures based on performance gaps"""
    underperforming = []

    for measure in measures:
        conway_mean = measure['summary']['conwayMean']
        national_mean = measure['summary']['nationalMean']

        if conway_mean is None or national_mean is None:
            continue

        # Skip if both are 0
        if conway_mean == 0 and national_mean == 0:
            continue

        # Calculate difference and percentage gap
        difference = abs(conway_mean - national_mean)

        if national_mean != 0:
            percentage_gap = ((conway_mean - national_mean) / national_mean) * 100
        else:
            percentage_gap = 0

        weight = measure['summary']['weight'] if measure['summary']['weight'] else 0

        underperforming.append({
            'measureGroup': measure['domain'],
            'measureId': measure['measureId'],
            'measureName': measure['measureName'],
            'conwayMean': conway_mean,
            'nationalMean': national_mean,
            'difference': difference,
            'percentageGap': percentage_gap,
            'weight': weight,
            'weightedImpact': difference * weight
        })

    # Sort by difference (descending)
    underperforming.sort(key=lambda x: x['difference'], reverse=True)

    return underperforming

def calculate_measure_group_scores(measures):
    """Calculate measure group scores by year and domain"""
    # Aggregate z-scores by domain and period
    domain_scores = {}

    for measure in measures:
        domain = measure['domain']

        if domain not in domain_scores:
            domain_scores[domain] = {
                '2023 Spring': [],
                '2023 Fall': [],
                '2024 Spring': [],
                '2024 Fall': [],
                '2025 Spring': [],
                '2025 Fall': []
            }

        for period, data in measure['periods'].items():
            zscore = data['conwayZScore']
            if zscore is not None and zscore != '*' and zscore != 'Not Included' and zscore != 0:
                try:
                    domain_scores[domain][period].append(float(zscore))
                except (ValueError, TypeError):
                    pass

    # Calculate average z-scores
    group_scores = []

    periods = ['2023 Spring', '2023 Fall', '2024 Spring', '2024 Fall', '2025 Spring', '2025 Fall']

    for period in periods:
        score_entry = {'year': period}

        for domain, scores in domain_scores.items():
            if scores[period]:
                avg_zscore = sum(scores[period]) / len(scores[period])
                score_entry[domain] = round(avg_zscore, 2)
            else:
                score_entry[domain] = 0

        group_scores.append(score_entry)

    return group_scores

def format_as_typescript(data, var_name, indent=0):
    """Format Python data as TypeScript/JavaScript code"""
    indent_str = '  ' * indent

    if isinstance(data, dict):
        lines = ['{']
        for key, value in data.items():
            formatted_value = format_as_typescript(value, var_name, indent + 1)
            # Handle key quoting
            if '-' in key or ' ' in key or key[0].isdigit():
                lines.append(f'{indent_str}  "{key}": {formatted_value},')
            else:
                lines.append(f'{indent_str}  {key}: {formatted_value},')
        lines.append(f'{indent_str}}}')
        return '\n'.join(lines)

    elif isinstance(data, list):
        if not data:
            return '[]'
        lines = ['[']
        for item in data:
            formatted_item = format_as_typescript(item, var_name, indent + 1)
            lines.append(f'{indent_str}  {formatted_item},')
        lines.append(f'{indent_str}]')
        return '\n'.join(lines)

    elif isinstance(data, str):
        # Escape quotes in strings
        escaped = data.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n')
        return f'"{escaped}"'

    elif isinstance(data, bool):
        return 'true' if data else 'false'

    elif data is None:
        return 'null'

    else:
        return str(data)

def generate_typescript_output(measures):
    """Generate TypeScript code for all data structures"""
    output = []

    # Header comment
    output.append('// Auto-generated from Excel data')
    output.append('// Source: Conway SC Leapfrog Historical Measure Analysis 2023-2025 incl Fall 2025.xlsx')
    output.append('// Generated: ' + datetime.now().strftime('%a %b %d %H:%M:%S %Z %Y'))
    output.append('')

    # Generate scatterPlotData
    scatter_data = generate_scatter_plot_data(measures)
    output.append('const scatterPlotData = ' + format_as_typescript(scatter_data, 'scatterPlotData'))
    output.append('')

    # Generate underperformingMeasuresData
    underperforming_data = calculate_underperforming_measures(measures)
    output.append('const underperformingMeasuresData = ' + format_as_typescript(underperforming_data, 'underperformingMeasuresData'))
    output.append('')

    # Generate measureGroupScoresData
    group_scores = calculate_measure_group_scores(measures)
    output.append('const measureGroupScoresData = ' + format_as_typescript(group_scores, 'measureGroupScoresData'))
    output.append('')

    # Generate performanceData
    performance_data = generate_performance_data(measures)
    output.append('const performanceData = ' + format_as_typescript(performance_data, 'performanceData'))
    output.append('')

    # Generate measureData
    measure_data = generate_measure_data(measures)
    output.append('const measureData = ' + format_as_typescript(measure_data, 'measureData'))
    output.append('')

    # Add export statement
    output.append('export { scatterPlotData, underperformingMeasuresData, measureGroupScoresData, performanceData, measureData }')
    output.append('')

    return '\n'.join(output)

def main():
    print('Parsing Excel file...')
    measures = parse_excel()
    print(f'Found {len(measures)} measures')

    print('\nGenerating TypeScript data structures...')
    ts_output = generate_typescript_output(measures)

    # Save to file
    output_file = '/Users/heilman/Desktop/Healthcare Dashboard/generated-data.ts'
    with open(output_file, 'w') as f:
        f.write(ts_output)

    print(f'\n✅ Data generated successfully!')
    print(f'📄 Output saved to: {output_file}')
    print(f'\n📊 Summary:')
    print(f'  - Total measures: {len(measures)}')

    # Count by domain
    domains = {}
    for m in measures:
        domain = m['domain']
        domains[domain] = domains.get(domain, 0) + 1

    for domain, count in domains.items():
        print(f'  - {domain}: {count} measures')

    # Also save raw JSON for inspection
    json_file = '/Users/heilman/Desktop/Healthcare Dashboard/generated-data.json'
    with open(json_file, 'w') as f:
        json.dump(measures, f, indent=2)
    print(f'\n📋 Raw JSON saved to: {json_file}')

if __name__ == '__main__':
    main()
